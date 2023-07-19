import openpyxl

wb = openpyxl.load_workbook("test.xlsx")
ws = wb.active
print('running openpyxl test')
assert ws['A2'].value == 42
assert ws['B2'].value == 'fish'

